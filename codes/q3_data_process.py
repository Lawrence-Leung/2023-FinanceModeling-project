'''
    第3题 数据处理代码
    2023/11/5
'''
import os
import json
import numpy as np
import openpyxl

# 函数：读取Excel文件并获取ETF数据
def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    etf_data = {}
    for column in range(2, sheet.max_column + 1, 5):
        etf_code = sheet.cell(row=1, column=column).value
        etf_data[etf_code] = []

        for row in range(4, sheet.max_row + 1):
            daily_data = [
                sheet.cell(row=row, column=column + i).value
                for i in range(5)
            ]
            etf_data[etf_code].append(daily_data)

    return etf_data
# 函数：将数据保存为numpy矩阵
def save_to_numpy(data_dict, folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    etf_code_map = {}
    for n, (etf_code, data) in enumerate(data_dict.items(), start=1):
        np_data = np.array(data)
        file_path = os.path.join(folder_path, f"{n}.npy")
        np.save(file_path, np_data)
        etf_code_map[etf_code] = n

    return etf_code_map
# 函数：将ETF代码和编号对应关系保存为json
def save_code_map(code_map, file_path):
    with open(file_path, 'w') as json_file:
        json.dump(code_map, json_file)
# 主执行函数
def process_etf_data(excel_file_path, output_folder_path, json_file_path):
    # 读取Excel文件
    etf_data = read_excel(excel_file_path)

    # 将ETF数据保存为numpy矩阵，并获取ETF代码与文件编号的映射
    code_map = save_to_numpy(etf_data, output_folder_path)

    # 将ETF代码与文件编号的映射保存为json文件
    save_code_map(code_map, json_file_path)

    return etf_data

#----------------------------------------------------------------------------
# 功能代码1
'''
# 使用例子
excel_file_path = 'domestic_options_data.xlsx'  # 将这里改为你的.xlsx文件路径
output_folder_path = 'domestic_options_data'                   # numpy矩阵将会保存在这个新文件夹中
json_file_path = 'domestic_options_code_map.json'              # ETF代码与文件编号对应关系的json文件路径

# 调用主执行函数
etf = process_etf_data(excel_file_path, output_folder_path, json_file_path)
'''

# 功能代码2
def g(X, Y, market_delta_all):
    # 验证输入的时间序列X和Y的长度是否相等
    if len(X) != len(Y):
        raise ValueError("The lengths of X and Y must be equal.")

    # 验证market_delta_all的长度是否为X和Y的长度减8
    if len(market_delta_all) != len(X) - 8:
        raise ValueError("The length of market_delta_all must be the length of X (or Y) minus 8.")

    # 定义结果列表
    results = []

    # 从t=6开始，直到n-3（因为我们需要包括n-3，所以循环到n-2）
    for t in range(6, len(X) - 2):
        # 从各个输入数组中获取相应的值
        value_X = X[t]
        value_Y = Y[t]
        value_market_delta = market_delta_all[t - 6]

        # 将这些值组合成一个元组并添加到结果列表中
        results.append((value_X, value_Y, value_market_delta))

    return results


# 测试函数
# 假设X, Y, market_delta_all都是numpy数组，且长度符合要求
# X, Y的长度为n, market_delta_all的长度为n-8
# 下面仅为示例数据
n = 20  # 假定n为20
X = np.random.rand(n)  # 生成一个长度为n的随机数组
Y = np.random.rand(n)  # 生成另一个长度为n的随机数组
market_delta_all = np.random.rand(n - 8)  # 生成第三个数组，长度为n-8

# 调用函数
result = g(X, Y, market_delta_all)

# 打印结果
print(result)
